import flet as ft
import os
import subprocess
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.styles.protection import Protection
from openpyxl.utils import get_column_letter
from Libs.Public.ui import configure_main_window, go_to_login
from Libs.Public.utils import create_drag_area, create_drawer, show_snackbar
from Libs.Data.sql_server_config import initialize_sql_server

def handle_change(e, page: ft.Page):
    from Libs.Public.menu import menu_page
    from Libs.Scripts.scripts import scripts_page
    from Libs.Wiki.wiki import wiki_page
    from Libs.Technical.technical import technical_page
    from Libs.Movdesk.movdesk import movdesk_page
    from Config.settings import settings_page

    page.clean()
    
    page_map = {
        0: menu_page,
        1: scripts_page,
        2: wiki_page,
        3: technical_page,
        4: migracao_page,
        5: movdesk_page,
        6: settings_page,
    }

    selected_index = e.control.selected_index
    if selected_index in page_map:
        page_map[selected_index](page)
    elif selected_index == 7:
        go_to_login(page)

    page.close(e.control)

def migracao_page(page: ft.Page):
    configure_main_window(page)
    page.title = "Menu Migração"
    page.window.title_bar_hidden = True
    page.window.maximizable = False
    page.window.resizable = False
    page.theme_mode = 'Dark'

    drawer = create_drawer(page)
    drawer.selected_index = 4
    drawer.on_change = lambda e: handle_change(e, page)
    
    drag_area = create_drag_area(page, drawer)

    botoes_container = ft.Column(
        controls=[
            ft.ElevatedButton("Download Planilha", on_click=baixar_planilha),
            ft.ElevatedButton("Upload Planilha", on_click=upload_planilha_sql),
            ft.ElevatedButton("Validações", on_click=executar_validacoes),
            ft.ElevatedButton("Correções", on_click=executar_correcoes),
        ],
        alignment=ft.MainAxisAlignment.CENTER,
        horizontal_alignment=ft.CrossAxisAlignment.CENTER
    )

    main_container = ft.Container(
        content=ft.Column(
            controls=[drag_area, botoes_container],
            expand=True,
            alignment=ft.MainAxisAlignment.CENTER,
            horizontal_alignment=ft.CrossAxisAlignment.CENTER
        ),
        padding=ft.padding.all(0),
        margin=ft.Margin(left=0, right=0, top=0, bottom=0),
    )

    page.add(main_container)
    page.update()

def baixar_planilha(e):
    tipos_planilhas = {
        "Clientes": ["ID", "Nome", "CNPJ", "Endereço"],
        "Produtos": ["ID", "Descrição", "Preço", "Quantidade"],
        "Vendas": ["ID", "Cliente", "Produto", "Quantidade", "Data"]
    }

    # Cria o conteúdo do diálogo com os botões para cada tipo de planilha
    dialog_content = ft.Column(
        controls=[
            ft.ElevatedButton(tipo, on_click=lambda e, tipo=tipo: criar_planilha(tipo, tipos_planilhas[tipo], e))
            for tipo in tipos_planilhas
        ]
    )

    # Cria o diálogo que permite ao usuário escolher o tipo de planilha
    dialog = ft.AlertDialog(
        title=ft.Text("Escolha o tipo de planilha para download:"),
        content=dialog_content,
        actions=[
            ft.TextButton("Fechar", on_click=lambda e: dialog.close())  # Fechar o diálogo
        ]
    )

    # Adiciona o diálogo à sobreposição da página
    e.page.overlay.append(dialog)
    
    # Abre o diálogo
    dialog.open = True
    e.page.update()  # Atualiza a página para refletir a mudança

def criar_planilha(tipo, titulos, e):
    # Defina o diretório onde você deseja salvar o arquivo
    diretorio_salvamento = os.path.join(os.path.dirname(__file__), "Excel")
    nome_arquivo = f"{tipo}.xlsx"
    caminho_arquivo = os.path.join(diretorio_salvamento, nome_arquivo)  # Cria o caminho completo

    wb = Workbook()  # Cria uma nova pasta de trabalho
    ws = wb.active
    ws.title = tipo  # Define o título da planilha

    # Adiciona os títulos das colunas à planilha
    for col, titulo in enumerate(titulos, start=1):
        cell = ws.cell(row=1, column=col, value=titulo)
        # Formata as células de título
        cell.protection = Protection(locked=True)
        cell.font = Font(bold=True, color="001219", size=14)
        cell.fill = PatternFill(start_color="00b4d8", end_color="00b4d8", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col)].width = 20  # Ajusta a largura da coluna

    # Desbloqueia as células a partir da segunda linha para que possam ser editadas
    for row in ws.iter_rows(min_row=2, max_col=len(titulos), max_row=100):
        for cell in row:
            cell.protection = Protection(locked=False)

    # Protege a planilha com uma senha
    ws.protection.sheet = True
    ws.protection.password = "senha_protecao"
    
    # Salva a planilha no caminho especificado
    wb.save(caminho_arquivo)
    print(f"Planilha '{caminho_arquivo}' criada com sucesso e protegida.")

    # Abre o arquivo no explorador de arquivos
    if os.path.exists(caminho_arquivo):
        subprocess.Popen(f'explorer "{caminho_arquivo}"')
        show_snackbar(e.page, f"Planilha '{tipo}' criada e aberta com sucesso.", is_error=False)
    else:
        show_snackbar(e.page, "A planilha não foi encontrada.", is_error=True)

def upload_planilha_sql(e):
    folder_picker = ft.FolderPicker(on_result=lambda e: processar_upload(e.path, e.page))
    e.page.overlay.append(folder_picker)
    folder_picker.pick()

def processar_upload(folder_path, page):
    if folder_path is None:
        show_snackbar(page, "Nenhuma pasta selecionada.", is_error=True)
        return

    try:
        connection = initialize_sql_server()
        if connection is None:
            show_snackbar(page, "Erro ao conectar ao banco de dados.", is_error=True)
            return

        for file_name in os.listdir(folder_path):
            if file_name.endswith('.xlsx'):
                file_path = os.path.join(folder_path, file_name)
                df = pd.read_excel(file_path)
                table_name = os.path.splitext(file_name)[0]
                df.to_sql(table_name, connection, if_exists="replace", index=False)
                print(f"Arquivo '{table_name}' enviado para o SQL Server.")
        
        show_snackbar(page, "Arquivos enviados com sucesso.", is_error=False)
    except Exception as e:
        show_snackbar(page, f"Erro ao enviar os arquivos: {str(e)}", is_error=True)

def executar_validacoes(e):
    print("Executando validações...")  

def executar_correcoes(e):
    print("Executando correções...")  
